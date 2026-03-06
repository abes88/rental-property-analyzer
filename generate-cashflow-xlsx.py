#!/usr/bin/env python3
"""
Rental Property Cash Flow Spreadsheet Generator

Generates a formatted .xlsx workbook with:
- An Inputs section with adjustable knobs (down payment %, interest rate, etc.)
- A Cash Flow section with Bear / Base / Bull rent scenario columns
- Yellow-highlighted cells for unknown values that need manual entry
- Live Excel formulas so changing any input auto-recalculates everything

Usage:
    python3 generate-cashflow-xlsx.py property-data.json output.xlsx

The JSON file should contain:
{
    "address": "123 Main St, City, ST 12345",
    "date": "2026-03-05",
    "purchase_price": 357400,
    "down_payment_pct": 0.25,
    "interest_rate": 0.07,
    "loan_term_years": 30,
    "closing_cost_pct": 0.02,
    "vacancy_loss_rate": 0.08,
    "mgmt_rate": 0.08,
    "property_tax_annual": 4074,
    "insurance_annual": null,
    "hoa_monthly": 831,
    "repairs_maintenance_annual": null,
    "gross_rent_bear": 2520,
    "gross_rent_base": 2800,
    "gross_rent_bull": 3200
}

Set any value to null to leave it blank (yellow-highlighted) in the spreadsheet.
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
RESULT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)
NORMAL_FONT = Font(size=11)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

DOLLAR_FMT = '$#,##0'
DOLLAR_CENTS_FMT = '$#,##0.00'
PCT_FMT = '0.00%'


def apply_style(cell, font=None, fill=None, alignment=None, number_format=None):
    """Apply formatting to a cell."""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format
    cell.border = THIN_BORDER


def write_cell(ws, row, col, value, font=NORMAL_FONT, fill=None, fmt=None):
    """Write a value to a cell with formatting. Yellow-fill if value is None (blank)."""
    cell = ws.cell(row=row, column=col)
    if value is None:
        cell.value = None
        apply_style(cell, font=font, fill=YELLOW_FILL, number_format=fmt)
    else:
        cell.value = value
        actual_fill = fill if fill else None
        apply_style(cell, font=font, fill=actual_fill, number_format=fmt)
    return cell


def write_formula(ws, row, col, formula, font=NORMAL_FONT, fill=None, fmt=None):
    """Write a formula to a cell with formatting."""
    cell = ws.cell(row=row, column=col)
    cell.value = formula
    apply_style(cell, font=font, fill=fill, number_format=fmt)
    return cell


def generate_workbook(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow Analysis"

    # Column widths: A=labels, B=Bear, C=Base, D=Bull
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    row = 1

    # -----------------------------------------------------------------------
    # Title
    # -----------------------------------------------------------------------
    ws.merge_cells('A1:D1')
    title_cell = ws.cell(row=1, column=1, value="RENTAL PROPERTY CASH FLOW ANALYSIS")
    apply_style(title_cell, font=TITLE_FONT, fill=HEADER_FILL, alignment=Alignment(horizontal="center"))
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    for col in range(2, 5):
        apply_style(ws.cell(row=1, column=col), fill=HEADER_FILL)

    row = 2
    ws.merge_cells('A2:D2')
    addr_cell = ws.cell(row=2, column=1, value=data.get("address", ""))
    apply_style(addr_cell, font=BOLD_FONT, alignment=Alignment(horizontal="center"))
    for col in range(2, 5):
        apply_style(ws.cell(row=2, column=col))

    row = 3
    ws.merge_cells('A3:D3')
    date_cell = ws.cell(row=3, column=1, value=f"Analysis Date: {data.get('date', '')}")
    apply_style(date_cell, font=NORMAL_FONT, alignment=Alignment(horizontal="center"))
    for col in range(2, 5):
        apply_style(ws.cell(row=3, column=col))

    # -----------------------------------------------------------------------
    # INPUTS section
    # -----------------------------------------------------------------------
    row = 5
    for col in range(1, 5):
        c = ws.cell(row=row, column=col)
        if col == 1:
            c.value = "INPUTS"
        apply_style(c, font=SECTION_FONT, fill=SECTION_FILL)

    # Helper to write an input row: label in A, value in B
    def input_row(r, label, value, fmt=DOLLAR_FMT):
        write_cell(ws, r, 1, label, font=NORMAL_FONT)
        write_cell(ws, r, 2, value, font=NORMAL_FONT, fmt=fmt)
        # Blank C and D for clean look
        for col in [3, 4]:
            apply_style(ws.cell(row=r, column=col))

    row = 6
    input_row(row, "Purchase Price", data.get("purchase_price"), DOLLAR_FMT)
    # Track key cell references for formulas
    CELL_PRICE = "B6"

    row = 7
    input_row(row, "Down Payment %", data.get("down_payment_pct"), PCT_FMT)
    CELL_DP_PCT = "B7"

    row = 8
    input_row(row, "Interest Rate (Annual)", data.get("interest_rate"), PCT_FMT)
    CELL_RATE = "B8"

    row = 9
    input_row(row, "Loan Term (Years)", data.get("loan_term_years", 30), '0')
    CELL_TERM = "B9"

    row = 10
    input_row(row, "Closing Costs %", data.get("closing_cost_pct", 0.02), PCT_FMT)
    CELL_CLOSING_PCT = "B10"

    row = 11
    input_row(row, "Vacancy & Loss Rate", data.get("vacancy_loss_rate", 0.08), PCT_FMT)
    CELL_VACANCY = "B11"

    row = 12
    input_row(row, "Property Management Rate", data.get("mgmt_rate", 0.08), PCT_FMT)
    CELL_MGMT = "B12"

    # Blank separator
    row = 13
    for col in range(1, 5):
        apply_style(ws.cell(row=row, column=col))

    # -----------------------------------------------------------------------
    # CALCULATED INPUTS (formulas derived from above)
    # -----------------------------------------------------------------------
    row = 14
    for col in range(1, 5):
        c = ws.cell(row=row, column=col)
        if col == 1:
            c.value = "LOAN DETAILS"
        apply_style(c, font=SECTION_FONT, fill=SECTION_FILL)

    row = 15
    write_cell(ws, row, 1, "Down Payment", font=NORMAL_FONT)
    write_formula(ws, row, 2, f"={CELL_PRICE}*{CELL_DP_PCT}", fmt=DOLLAR_FMT)
    for col in [3, 4]:
        apply_style(ws.cell(row=row, column=col))
    CELL_DP = "B15"

    row = 16
    write_cell(ws, row, 1, "Loan Amount", font=NORMAL_FONT)
    write_formula(ws, row, 2, f"={CELL_PRICE}-{CELL_DP}", fmt=DOLLAR_FMT)
    for col in [3, 4]:
        apply_style(ws.cell(row=row, column=col))
    CELL_LOAN = "B16"

    row = 17
    write_cell(ws, row, 1, "Closing Costs", font=NORMAL_FONT)
    write_formula(ws, row, 2, f"={CELL_PRICE}*{CELL_CLOSING_PCT}", fmt=DOLLAR_FMT)
    for col in [3, 4]:
        apply_style(ws.cell(row=row, column=col))
    CELL_CLOSING = "B17"

    row = 18
    write_cell(ws, row, 1, "Total Cash Invested", font=BOLD_FONT)
    write_formula(ws, row, 2, f"={CELL_DP}+{CELL_CLOSING}", font=BOLD_FONT, fmt=DOLLAR_FMT)
    for col in [3, 4]:
        apply_style(ws.cell(row=row, column=col))
    CELL_CASH_INVESTED = "B18"

    row = 19
    write_cell(ws, row, 1, "Monthly P&I", font=BOLD_FONT)
    # PMT formula: L * [r(1+r)^n] / [(1+r)^n - 1], where r=monthly rate, n=months
    # Excel-style: uses the PMT-equivalent formula since openpyxl supports standard Excel formulas
    # -PMT(rate/12, term*12, loan_amount) gives positive monthly payment
    write_formula(
        ws, row, 2,
        f"=-PMT({CELL_RATE}/12,{CELL_TERM}*12,{CELL_LOAN})",
        font=BOLD_FONT, fmt=DOLLAR_FMT
    )
    for col in [3, 4]:
        apply_style(ws.cell(row=row, column=col))
    CELL_PI = "B19"

    # -----------------------------------------------------------------------
    # MONTHLY EXPENSE INPUTS
    # -----------------------------------------------------------------------
    row = 21
    for col in range(1, 5):
        c = ws.cell(row=row, column=col)
        if col == 1:
            c.value = "MONTHLY EXPENSE INPUTS"
        apply_style(c, font=SECTION_FONT, fill=SECTION_FILL)

    row = 22
    write_cell(ws, row, 1, "Property Tax (Annual)", font=NORMAL_FONT)
    write_cell(ws, row, 2, data.get("property_tax_annual"), font=NORMAL_FONT, fmt=DOLLAR_FMT)
    for col in [3, 4]:
        apply_style(ws.cell(row=row, column=col))
    CELL_TAX_ANNUAL = "B22"

    row = 23
    write_cell(ws, row, 1, "Insurance (Annual)", font=NORMAL_FONT)
    write_cell(ws, row, 2, data.get("insurance_annual"), font=NORMAL_FONT, fmt=DOLLAR_FMT)
    for col in [3, 4]:
        apply_style(ws.cell(row=row, column=col))
    CELL_INS_ANNUAL = "B23"

    row = 24
    write_cell(ws, row, 1, "HOA (Monthly)", font=NORMAL_FONT)
    write_cell(ws, row, 2, data.get("hoa_monthly"), font=NORMAL_FONT, fmt=DOLLAR_FMT)
    for col in [3, 4]:
        apply_style(ws.cell(row=row, column=col))
    CELL_HOA = "B24"

    row = 25
    write_cell(ws, row, 1, "Repairs & Maintenance (Annual)", font=NORMAL_FONT)
    write_cell(ws, row, 2, data.get("repairs_maintenance_annual"), font=NORMAL_FONT, fmt=DOLLAR_FMT)
    for col in [3, 4]:
        apply_style(ws.cell(row=row, column=col))
    CELL_MAINT_ANNUAL = "B25"

    # -----------------------------------------------------------------------
    # MONTHLY CASH FLOW — Bear / Base / Bull
    # -----------------------------------------------------------------------
    row = 27
    for col in range(1, 5):
        c = ws.cell(row=row, column=col)
        vals = {1: "MONTHLY CASH FLOW", 2: "Bear", 3: "Base", 4: "Bull"}
        c.value = vals.get(col)
        apply_style(c, font=SECTION_FONT, fill=SECTION_FILL)

    # --- Gross Monthly Rent ---
    row = 28
    write_cell(ws, row, 1, "Gross Monthly Rent", font=NORMAL_FONT)
    write_cell(ws, row, 2, data.get("gross_rent_bear"), font=NORMAL_FONT, fmt=DOLLAR_FMT)
    write_cell(ws, row, 3, data.get("gross_rent_base"), font=NORMAL_FONT, fmt=DOLLAR_FMT)
    write_cell(ws, row, 4, data.get("gross_rent_bull"), font=NORMAL_FONT, fmt=DOLLAR_FMT)
    ROW_RENT = 28

    # --- Vacancy & Loss ---
    row = 29
    write_cell(ws, row, 1, "Less: Vacancy & Loss (8%)", font=NORMAL_FONT)
    for col in [2, 3, 4]:
        cl = get_column_letter(col)
        write_formula(ws, row, col, f"=-{cl}{ROW_RENT}*{CELL_VACANCY}", fmt=DOLLAR_FMT)
    ROW_VAC = 29

    # --- Effective Gross Income ---
    row = 30
    write_cell(ws, row, 1, "Effective Gross Income", font=BOLD_FONT)
    for col in [2, 3, 4]:
        cl = get_column_letter(col)
        write_formula(ws, row, col, f"={cl}{ROW_RENT}+{cl}{ROW_VAC}", font=BOLD_FONT, fmt=DOLLAR_FMT)
    ROW_EGI = 30

    # --- Operating Expenses (all as negative values) ---
    row = 31
    write_cell(ws, row, 1, "Less: Property Tax", font=NORMAL_FONT)
    for col in [2, 3, 4]:
        write_formula(ws, row, col, f"=-{CELL_TAX_ANNUAL}/12", fmt=DOLLAR_FMT)
    ROW_TAX = 31

    row = 32
    write_cell(ws, row, 1, "Less: Insurance", font=NORMAL_FONT)
    for col in [2, 3, 4]:
        write_formula(ws, row, col, f"=-{CELL_INS_ANNUAL}/12", fmt=DOLLAR_FMT)
    ROW_INS = 32

    row = 33
    write_cell(ws, row, 1, "Less: HOA", font=NORMAL_FONT)
    for col in [2, 3, 4]:
        write_formula(ws, row, col, f"=-{CELL_HOA}", fmt=DOLLAR_FMT)
    ROW_HOA = 33

    row = 34
    write_cell(ws, row, 1, "Less: Repairs & Maintenance", font=NORMAL_FONT)
    for col in [2, 3, 4]:
        write_formula(ws, row, col, f"=-{CELL_MAINT_ANNUAL}/12", fmt=DOLLAR_FMT)
    ROW_MAINT = 34

    row = 35
    write_cell(ws, row, 1, "Less: Property Management (8%)", font=NORMAL_FONT)
    for col in [2, 3, 4]:
        cl = get_column_letter(col)
        write_formula(ws, row, col, f"=-{cl}{ROW_EGI}*{CELL_MGMT}", fmt=DOLLAR_FMT)
    ROW_PM = 35

    # --- Net Operating Income ---
    row = 37
    for col in range(1, 5):
        apply_style(ws.cell(row=36, column=col))  # blank separator row
    write_cell(ws, row, 1, "NET OPERATING INCOME (Monthly)", font=BOLD_FONT)
    for col in [2, 3, 4]:
        cl = get_column_letter(col)
        write_formula(
            ws, row, col,
            f"={cl}{ROW_EGI}+{cl}{ROW_TAX}+{cl}{ROW_INS}+{cl}{ROW_HOA}+{cl}{ROW_MAINT}+{cl}{ROW_PM}",
            font=BOLD_FONT, fill=RESULT_FILL, fmt=DOLLAR_FMT
        )
    ROW_NOI = 37

    # --- Debt Service ---
    row = 38
    write_cell(ws, row, 1, "Less: Debt Service (P&I)", font=NORMAL_FONT)
    for col in [2, 3, 4]:
        write_formula(ws, row, col, f"=-{CELL_PI}", fmt=DOLLAR_FMT)
    ROW_DEBT = 38

    # --- Cash Flow After Debt Service ---
    row = 39
    write_cell(ws, row, 1, "CASH FLOW AFTER DEBT SERVICE", font=BOLD_FONT)
    for col in [2, 3, 4]:
        cl = get_column_letter(col)
        write_formula(
            ws, row, col,
            f"={cl}{ROW_NOI}+{cl}{ROW_DEBT}",
            font=BOLD_FONT, fill=RESULT_FILL, fmt=DOLLAR_FMT
        )
    ROW_CF = 39

    # -----------------------------------------------------------------------
    # ANNUAL SUMMARY
    # -----------------------------------------------------------------------
    row = 41
    for col in range(1, 5):
        c = ws.cell(row=row, column=col)
        if col == 1:
            c.value = "ANNUAL SUMMARY"
        apply_style(c, font=SECTION_FONT, fill=SECTION_FILL)

    row = 42
    write_cell(ws, row, 1, "Annual NOI", font=NORMAL_FONT)
    for col in [2, 3, 4]:
        cl = get_column_letter(col)
        write_formula(ws, row, col, f"={cl}{ROW_NOI}*12", fmt=DOLLAR_FMT)
    ROW_ANNUAL_NOI = 42

    row = 43
    write_cell(ws, row, 1, "Annual Debt Service", font=NORMAL_FONT)
    for col in [2, 3, 4]:
        write_formula(ws, row, col, f"={CELL_PI}*12", fmt=DOLLAR_FMT)
    ROW_ANNUAL_DEBT = 43

    row = 44
    write_cell(ws, row, 1, "Annual Cash Flow", font=BOLD_FONT)
    for col in [2, 3, 4]:
        cl = get_column_letter(col)
        write_formula(ws, row, col, f"={cl}{ROW_CF}*12", font=BOLD_FONT, fill=RESULT_FILL, fmt=DOLLAR_FMT)
    ROW_ANNUAL_CF = 44

    row = 45
    write_cell(ws, row, 1, "Cash on Cash Return", font=BOLD_FONT)
    for col in [2, 3, 4]:
        cl = get_column_letter(col)
        write_formula(
            ws, row, col,
            f"=IF({CELL_CASH_INVESTED}=0,0,{cl}{ROW_ANNUAL_CF}/{CELL_CASH_INVESTED})",
            font=BOLD_FONT, fill=RESULT_FILL, fmt=PCT_FMT
        )

    # -----------------------------------------------------------------------
    # Print settings
    # -----------------------------------------------------------------------
    ws.sheet_properties.pageSetUpPr = None
    ws.print_area = "A1:D45"

    wb.save(output_path)
    print(f"Saved: {output_path}")


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 generate-cashflow-xlsx.py <input.json> <output.xlsx>")
        sys.exit(1)

    with open(sys.argv[1], "r") as f:
        data = json.load(f)

    generate_workbook(data, sys.argv[2])


if __name__ == "__main__":
    main()
